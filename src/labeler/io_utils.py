from __future__ import annotations

import csv
import json
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class InputFileFormatError(ValueError):
    def __init__(self, message: str, can_try_csv: bool = False):
        super().__init__(message)
        self.can_try_csv = can_try_csv


@dataclass
class DialogueRecord:
    session_id: str
    text: str
    row_index: int = 0
    annotation_key: str = ''


@dataclass
class InputFileStatus:
    ok: bool
    warning: str = ''


_ALLOWED_EXTENSIONS = {'.xlsx', '.xlsm', '.csv'}


def list_input_files(input_dir):
    files = []
    for path in sorted(Path(input_dir).iterdir()):
        if not path.is_file():
            continue
        if path.name.startswith('~$'):
            continue
        if path.suffix.lower() in _ALLOWED_EXTENSIONS:
            files.append(path)
    return files


def inspect_input_file(file_path):
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    if suffix in ('.xlsx', '.xlsm'):
        if not zipfile.is_zipfile(str(file_path)):
            return InputFileStatus(
                ok=False,
                warning=(
                    'Расширение Excel, но файл не является корректным XLSX/XLSM. '
                    'Возможно, это CSV, сохранённый с неправильным расширением.'
                ),
            )
    return InputFileStatus(ok=True, warning='')


def split_dialogue_text(text):
    if text is None:
        return []
    text = str(text)
    parts = [part.strip() for part in text.split('//')]
    return [part for part in parts if part]


def _normalize_headers(headers):
    result = []
    for header in headers:
        result.append(str(header).strip() if header is not None else '')
    return result


def _assign_annotation_keys(records):
    counts = {}
    for record in records:
        sid = record.session_id.strip()
        counts[sid] = counts.get(sid, 0) + 1
    for record in records:
        sid = record.session_id.strip()
        if sid and counts.get(sid, 0) == 1:
            record.annotation_key = sid
        else:
            base = sid if sid else 'row'
            record.annotation_key = '{0}__row_{1}'.format(base, record.row_index)
    return records


def _load_from_excel(file_path):
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('Файл пустой.')
    headers = _normalize_headers(rows[0])
    try:
        session_idx = headers.index('session_id')
        text_idx = headers.index('text')
    except ValueError:
        raise ValueError('Во входном файле должны быть колонки session_id и text.')
    records = []
    for row_index, row in enumerate(rows[1:], start=2):
        session_id = '' if session_idx >= len(row) or row[session_idx] is None else str(row[session_idx]).strip()
        text = '' if text_idx >= len(row) or row[text_idx] is None else str(row[text_idx]).strip()
        if not session_id and not text:
            continue
        records.append(DialogueRecord(session_id=session_id, text=text, row_index=row_index))
    return _assign_annotation_keys(records)


def _load_from_csv(file_path):
    records = []
    with Path(file_path).open('r', encoding='utf-8-sig', newline='') as fh:
        reader = csv.DictReader(fh, delimiter=';')
        if not reader.fieldnames or 'session_id' not in reader.fieldnames or 'text' not in reader.fieldnames:
            raise ValueError('Во входном файле должны быть колонки session_id и text.')
        for idx, row in enumerate(reader, start=2):
            session_id = str(row.get('session_id', '')).strip()
            text = str(row.get('text', '')).strip()
            if not session_id and not text:
                continue
            records.append(DialogueRecord(session_id=session_id, text=text, row_index=idx))
    return _assign_annotation_keys(records)


def load_dialogues(file_path, treat_as_csv=False):
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    if treat_as_csv:
        records = _load_from_csv(file_path)
    elif suffix in ('.xlsx', '.xlsm'):
        if not zipfile.is_zipfile(str(file_path)):
            raise InputFileFormatError(
                (
                    'Файл имеет расширение {0}, но по содержимому не является Excel-файлом. '
                    'Возможно, это CSV, сохранённый с неправильным расширением.'
                ).format(suffix),
                can_try_csv=True,
            )
        records = _load_from_excel(file_path)
    elif suffix == '.csv':
        records = _load_from_csv(file_path)
    else:
        raise ValueError('Поддерживаются только .xlsx, .xlsm и .csv.')
    if not records:
        raise ValueError('Во входном файле не найдено ни одной записи.')
    return records, {'total_records': len(records), 'source_file': file_path.name}


def normalize_annotations_for_records(records, raw_annotations):
    if not isinstance(raw_annotations, dict):
        return {}
    result = {}
    used_old_keys = set()
    for record in records:
        if record.annotation_key in raw_annotations and isinstance(raw_annotations.get(record.annotation_key), dict):
            result[record.annotation_key] = raw_annotations.get(record.annotation_key)
            used_old_keys.add(record.annotation_key)
            continue
        if record.session_id in raw_annotations and isinstance(raw_annotations.get(record.session_id), dict) and record.session_id not in used_old_keys:
            result[record.annotation_key] = raw_annotations.get(record.session_id)
            used_old_keys.add(record.session_id)
    return result


def build_draft_path(output_dir, input_path, mode_id):
    drafts_dir = Path(output_dir) / '_drafts'
    drafts_dir.mkdir(parents=True, exist_ok=True)
    return drafts_dir / '{0}__{1}__draft.json'.format(Path(input_path).stem, mode_id)


def _safe_name_fragment(value):
    text = str(value or '').strip()
    if not text:
        return 'mode'
    allowed = []
    for ch in text:
        if ch.isalnum() or ch in ('-', '_'):
            allowed.append(ch)
        else:
            allowed.append('_')
    compact = ''.join(allowed)
    while '__' in compact:
        compact = compact.replace('__', '_')
    return compact.strip('_') or 'mode'


def build_output_path(output_dir, input_path, mode_id, completed_count=None, total_count=None, timestamp=None):
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    ts = timestamp or datetime.now().strftime('%Y_%m_%d_%H_%M')
    progress = ''
    if completed_count is not None and total_count is not None:
        progress = '__{0}_{1}'.format(int(completed_count), int(total_count))
    return Path(output_dir) / '{0}__{1}{2}__{3}.xlsx'.format(
        _safe_name_fragment(mode_id),
        ts,
        progress,
        _safe_name_fragment(Path(input_path).stem),
    )


def load_draft_json(draft_path):
    path = Path(draft_path)
    if not path.exists():
        return None
    try:
        with path.open('r', encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return None


def save_draft_json(draft_path, input_path, mode, annotations, current_index, annotator_name, filter_unlabeled):
    payload = {
        'source_file': Path(input_path).name,
        'mode_id': mode.get('id', ''),
        'mode_version': mode.get('version', ''),
        'current_index': int(current_index),
        'annotator_name': annotator_name,
        'filter_unlabeled': bool(filter_unlabeled),
        'annotations': annotations,
    }
    path = Path(draft_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8') as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


def _autosize_sheet(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            text = '' if value is None else str(value)
            widths[idx] = max(widths.get(idx, 0), min(80, len(text) + 2))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = max(12, width)


def save_results_excel(output_path, input_path, mode, records, annotations, annotator_name, summary=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'annotations'
    fields = mode.get('fields', [])
    meta_headers = ['confidence', 'needs_review', 'annotator_comment', 'completed', 'updated_at', 'annotator_name']
    headers = ['session_id', 'text'] + [field.get('key', '') for field in fields] + meta_headers
    ws.append(headers)
    for record in records:
        ann = annotations.get(record.annotation_key, {})
        row = [record.session_id, record.text]
        for field in fields:
            value = ann.get(field.get('key', ''), '')
            if isinstance(value, list):
                value = '; '.join([str(item) for item in value])
            row.append(value)
        row.extend([
            ann.get('confidence', ''),
            ann.get('needs_review', False),
            ann.get('annotator_comment', ''),
            ann.get('completed', False),
            ann.get('updated_at', ''),
            annotator_name,
        ])
        ws.append(row)
    _autosize_sheet(ws)

    stats = wb.create_sheet('stats')
    stats.append(['key', 'value'])
    stats.append(['source_file', Path(input_path).name])
    stats.append(['mode_id', mode.get('id', '')])
    stats.append(['mode_name', mode.get('name', '')])
    stats.append(['mode_version', mode.get('version', '')])
    stats.append(['annotator_name', annotator_name])
    summary = summary or {}
    for key in ['total_records', 'completed_records', 'remaining_records', 'needs_review_count']:
        stats.append([key, summary.get(key, '')])
    _autosize_sheet(stats)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
